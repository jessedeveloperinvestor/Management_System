#library to import the excel file
import openpyxl
#libraries to create the pdf file and add text to it
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
#library to get logo related information
from PIL import Image

#convert the font so it is compatible
pdfmetrics.registerFont(TTFont('Arial','Arial.ttf'))

#import the sheet from the excel file
wb = openpyxl.load_workbook('C:\\Users\\macbook\\Desktop\\ManagementSystem\\data.xlsx')
sheet = wb.get_sheet_by_name('servicos')

#import company's logo
im = Image.open('AO.png')
width, height = im.size
ratio = width/height
image_width = 400
image_height = int(image_width / ratio)

#Page information
page_width = 2156
page_height = 3050

#Invoice variables
company_name ='Auto Center Oliveira'
payment_terms = 'x'
contact_info = 'x'
margin = 100
month_year = 'February 2022'

#def function
def create_invoice():
    for i in range(2,3):
        #Reading values from excel file
        Cliente = sheet.cell(row = i, column = 2).value
        Numero_Orcamento = sheet.cell(row = i, column = 1).value
        Data_Inicial = sheet.cell(row = i, column = 13).value
        Data_Final = sheet.cell(row = i, column = 14).value
        Servico_1 = sheet.cell(row = i, column = 3).value
        Preco_1 = sheet.cell(row = i, column = 4).value
        Servico_2 = sheet.cell(row = i, column = 5).value
        Preco_2 = sheet.cell(row = i, column = 6).value
        Servico_3 = sheet.cell(row = i, column = 7).value
        Preco_3 = sheet.cell(row = i, column = 8).value
        Servico_4 = sheet.cell(row = i, column = 9).value
        Preco_4 = sheet.cell(row = i, column = 10).value
        Servico_5 = sheet.cell(row = i, column = 11).value
        Preco_5 = sheet.cell(row = i, column = 12).value

        Preco_1N = int(Preco_1)
        Preco_2N = int(Preco_2)
        Preco_3N = int(Preco_3)
        Preco_4N = int(Preco_4)
        Preco_5N = int(Preco_5)
        PrecoTotalSomadoList = []
        PrecoTotalSomadoList.append(Preco_1)
        PrecoTotalSomadoList.append(Preco_2)
        PrecoTotalSomadoList.append(Preco_3)
        PrecoTotalSomadoList.append(Preco_4)
        PrecoTotalSomadoList.append(Preco_5)
        sumprices = 0
        for num in PrecoTotalSomadoList:
            sumprices = sumprices +num
        PrecoTotalSomado = sumprices

        #Creating a pdf file and setting a naming convention
        c = canvas.Canvas(str(Numero_Orcamento) + '_' + str(Cliente) +'.pdf')
        c.setPageSize((page_width, page_height))

        #Drawing the image
        c.drawInlineImage("C:\\Users\\macbook\\Desktop\\ManagementSystem\\AO.png", page_width - image_width - margin,
                          page_height - image_height - margin,
                          image_width, image_height)

        #Invoice information
        c.setFont('Arial',80)
        text = 'NOTA DE SERVIÇOS'
        text_width = stringWidth(text,'Arial',80)
        c.drawString((page_width-text_width)/2, page_height - image_height - margin, text)
        y = page_height - image_height - margin*4
        x = 2*margin
        x2 = x + 550
        
        c.setFont('Arial', 45)
        c.drawString(x, y, 'Gerado por: ')
        c.drawString(x2,y, company_name)
        y -= margin
        
        c.drawString(x,y,'Cliente: ')
        c.drawString(x2,y,str(Cliente))
        y -= margin
        
        c.drawString(x,y,'Numero do Orçamento: ')
        c.drawString(x2,y, str(Numero_Orcamento))
        y -= margin
        
        Data_InicialF = str(Data_Inicial)[:-9]
        Data_FinalF = str(Data_Final)[:-9]

        c.drawString(x,y, 'Data Inicial: ')
        c.drawString(x2,y, str(Data_InicialF))
        y -= margin
        
        c.drawString(x,y,'Data Final: ')
        c.drawString(x2,y, str(Data_FinalF))
        y -= margin *2

        c.drawString(x,y,'Servico 1: ')
        c.drawString(x2,y, str(Servico_1))
        y-= margin
        
        c.drawString(x,y, 'Preço do Serviço 1: ')
        c.drawString(x2,y, 'R$ ' + str(Preco_1) + ',00')
        y -= margin

        c.drawString(x,y,'Servico 2: ')
        c.drawString(x2,y, str(Servico_2))
        y-= margin
        
        c.drawString(x,y, 'Preço do Serviço 2: ')
        c.drawString(x2,y, 'R$ ' + str(Preco_2) + ',00')
        y -= margin

        c.drawString(x,y,'Servico 3: ')
        c.drawString(x2,y, str(Servico_3))
        y-= margin
        
        c.drawString(x,y, 'Preço do Serviço 3: ')
        c.drawString(x2,y, 'R$ ' + str(Preco_3) + ',00')
        y -= margin

        c.drawString(x,y,'Servico 4: ')
        c.drawString(x2,y, str(Servico_4))
        y-= margin
        
        c.drawString(x,y, 'Preço do Serviço 4: ')
        c.drawString(x2,y, 'R$ ' + str(Preco_4) + ',00')
        y -= margin

        c.drawString(x,y,'Servico 5: ')
        c.drawString(x2,y, str(Servico_5))
        y-= margin
        
        c.drawString(x,y, 'Preço do Serviço 5: ')
        c.drawString(x2,y, 'R$ ' + str(Preco_5) + ',00')
        y -= margin

        c.drawString(x,y,'Preço Total: ')
        c.drawString(x2,y,'R$ ' + str(PrecoTotalSomado) + ',00')
        y -= margin*3
               
        c.drawString(x,y,'Obrigado pela escolha!')
        y -= margin
        c.drawString(x,y,'Fone: (11) 1234-5678')
        y -= margin
        c.drawString(x,y,'Sistemas empresariais e comerciais: https://jesse-leite-softwares.onrender.com')    

        #Saving the pdf file
        c.save()

create_invoice()
